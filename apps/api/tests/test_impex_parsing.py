"""Reading an uploaded table (issue #77): CSV, TSV, paste, Excel — and the guards.

Pure unit tests: no database, no tenant, no HTTP. Parsing is the layer that touches untrusted
bytes, so it is worth exercising directly rather than only through an import round-trip, where
a defence that silently never fires looks exactly like one that works.
"""

from __future__ import annotations

import datetime as dt
import io
import zipfile

import pytest
from openpyxl import Workbook

from app.core.impex import parsing
from app.core.impex.parsing import MAX_IMPORT_ROWS, parse_source
from app.errors import AppError


def _xlsx(
    rows: list[list[object]], *, sheets: dict[str, list[list[object]]] | None = None
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Klanten"
    for row in rows:
        sheet.append(row)
    for name, extra in (sheets or {}).items():
        other = workbook.create_sheet(name)
        for row in extra:
            other.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _error(raw: bytes, **kwargs: object) -> AppError:
    with pytest.raises(AppError) as caught:
        parse_source(raw, **kwargs)  # type: ignore[arg-type]
    return caught.value


# --------------------------------------------------------------------------- #
# Delimited text
# --------------------------------------------------------------------------- #
def test_tab_semicolon_and_comma_are_all_detected() -> None:
    """The tab case is the whole point of the paste box: a spreadsheet puts tabs on the
    clipboard, and the old sniffer (`,;` only) parsed such a paste as a single column."""
    expected = (["name", "city"], [["Acme", "Utrecht"], ["Beta", "Amsterdam"]])
    for delimiter in ("\t", ";", ",", "|"):
        body = f"name{delimiter}city\nAcme{delimiter}Utrecht\nBeta{delimiter}Amsterdam\n"
        table = parse_source(body.encode("utf-8"))
        assert (table.header, table.rows) == expected
        assert table.delimiter == delimiter
        assert table.source_format == "csv"


def test_a_single_column_paste_is_not_split() -> None:
    table = parse_source(b"name\nAcme\nBeta\n")
    assert table.header == ["name"]
    assert table.rows == [["Acme"], ["Beta"]]


def test_a_dutch_excel_csv_in_cp1252_is_read_not_rejected() -> None:
    """`utf-8` first, then `cp1252` — the encoding a Dutch Excel writes, which used to be an
    outright `invalid_file` and is the single most likely file an agency actually has."""
    raw = "name;city\nAcmé BV;Nijmegen\n".encode("cp1252")
    with pytest.raises(UnicodeDecodeError):
        raw.decode("utf-8")  # the file the old parser refused

    table = parse_source(raw)
    assert table.encoding == "cp1252"
    assert table.rows == [["Acmé BV", "Nijmegen"]]


def test_a_utf8_bom_is_stripped_from_the_first_header() -> None:
    table = parse_source("﻿name,city\nAcme,Utrecht\n".encode())
    assert table.header == ["name", "city"]


def test_rows_are_padded_to_one_width() -> None:
    """A short row is empty cells, never an IndexError downstream — which is what lets the
    mapping step address column 3 without every consumer re-checking the row length."""
    table = parse_source(b"name,city,phone\nAcme,Utrecht,06\nBeta\n")
    assert table.rows == [["Acme", "Utrecht", "06"], ["Beta", "", ""]]


def test_a_header_less_source_keeps_every_row_as_data() -> None:
    table = parse_source(b"Acme,Utrecht\nBeta,Amsterdam\n", has_header=False)
    assert table.header == []
    assert table.rows == [["Acme", "Utrecht"], ["Beta", "Amsterdam"]]


# --------------------------------------------------------------------------- #
# Excel
# --------------------------------------------------------------------------- #
def test_an_xlsx_parses_to_exactly_what_the_same_csv_does() -> None:
    """The engine downstream must not be able to tell the two apart."""
    csv_table = parse_source(b"name,city\nAcme,Utrecht\nBeta,Amsterdam\n")
    xlsx_table = parse_source(
        _xlsx([["name", "city"], ["Acme", "Utrecht"], ["Beta", "Amsterdam"]])
    )
    assert (xlsx_table.header, xlsx_table.rows) == (csv_table.header, csv_table.rows)
    assert xlsx_table.source_format == "xlsx"
    assert xlsx_table.sheet == "Klanten"


def test_excel_types_arrive_as_the_text_their_column_expects() -> None:
    """Excel has no text/number distinction: a client number is a float, a date a formatted
    float. `str()` on either produces something the column's own coercion then rejects —
    "1234.0" is not a client number and "2026-01-05 00:00:00" is not a date."""
    table = parse_source(
        _xlsx(
            [
                ["client_number", "postal_code", "start_date", "at", "active"],
                [1234, 3512.0, dt.datetime(2026, 1, 5), dt.datetime(2026, 1, 5, 9, 30), True],
            ]
        )
    )
    assert table.rows == [["1234", "3512", "2026-01-05", "2026-01-05 09:30", "true"]]


def test_a_named_sheet_is_read_and_an_unknown_one_is_an_error() -> None:
    raw = _xlsx(
        [["name"], ["Acme"]], sheets={"Archief": [["name"], ["Oud BV"]]}
    )
    assert parse_source(raw).sheets == ("Klanten", "Archief")
    assert parse_source(raw, sheet="Archief").rows == [["Oud BV"]]
    assert _error(raw, sheet="Ontbreekt").message_key == "impex.errors.sheet_not_found"


def test_an_uncalculated_formula_is_counted_not_silently_empty() -> None:
    """A workbook written by a library carries the formula but no cached result, and the
    cached result is what we read. One blank column that *looks* deliberate is the failure
    mode; the count is what lets the wizard say so."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["name", "full"])
    sheet.append(["Acme", "=A2&\" BV\""])
    buffer = io.BytesIO()
    workbook.save(buffer)

    table = parse_source(buffer.getvalue())
    assert table.rows == [["Acme", ""]]
    assert table.uncalculated_formulas == 1


def test_a_fully_populated_sheet_pays_nothing_for_the_formula_count() -> None:
    table = parse_source(_xlsx([["name", "city"], ["Acme", "Utrecht"]]))
    assert table.uncalculated_formulas == 0


def test_a_non_zip_file_is_a_clean_error_never_a_traceback() -> None:
    assert _error(b"PK\x03\x04not really a workbook").message_key == "impex.errors.invalid_xlsx"


def test_a_high_ratio_zip_is_refused_before_anything_is_decompressed(monkeypatch) -> None:
    """The guard reads the archive's **declared** sizes from the central directory, which is
    all a bomb has to lie about. Proven by making the parse itself explode: if the guard ran
    after `load_workbook`, this test would see that failure instead."""
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("xl/worksheets/sheet1.xml", b"\0" * (2 * 1024 * 1024))
    bomb = buffer.getvalue()
    assert len(bomb) < 64 * 1024  # compresses ~1000:1 — the ratio is the tell

    def _explode(*args: object, **kwargs: object) -> None:
        raise AssertionError("the workbook must never be opened")

    monkeypatch.setattr("openpyxl.load_workbook", _explode)
    error = _error(bomb)
    assert error.message_key == "impex.errors.file_too_large"
    assert error.status_code == 413


def test_too_many_members_is_refused() -> None:
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w") as archive:
        for index in range(parsing.MAX_XLSX_MEMBERS + 1):
            archive.writestr(f"part{index}.xml", b"x")
    assert _error(buffer.getvalue()).message_key == "impex.errors.invalid_xlsx"


# --------------------------------------------------------------------------- #
# Caps
# --------------------------------------------------------------------------- #
def test_a_too_wide_sheet_is_refused_rather_than_trimmed() -> None:
    width = parsing.MAX_COLUMNS + 5
    csv_body = ",".join(f"c{i}" for i in range(width)) + "\n" + ",".join("x" * width) + "\n"
    assert _error(csv_body.encode()).message_key == "impex.errors.too_many_columns"
    assert (
        _error(_xlsx([[f"c{i}" for i in range(width)], ["x"] * width])).message_key
        == "impex.errors.too_many_columns"
    )


def test_one_row_over_the_cap_is_a_413_never_a_silent_truncation() -> None:
    body = "name\n" + "".join(f"Acme {i}\n" for i in range(MAX_IMPORT_ROWS + 1))
    error = _error(body.encode())
    assert (error.message_key, error.status_code) == ("impex.errors.too_many_rows", 413)

    at_cap = "name\n" + "".join(f"Acme {i}\n" for i in range(MAX_IMPORT_ROWS))
    assert len(parse_source(at_cap.encode()).rows) == MAX_IMPORT_ROWS


def test_a_paste_has_its_own_smaller_cap() -> None:
    """Starlette truncates a non-file part past 1 MiB, so a larger limit could never be
    enforced — the check would be reading bytes that had already been cut."""
    # Rows wide enough that the row cap is *not* what trips: 2000 of these clear 1 MiB, so
    # the two limits are told apart rather than one masking the other.
    row = "Acme BV," + "x" * 600 + ",Utrecht\n"
    body = ("name,notes,city\n" + row * MAX_IMPORT_ROWS).encode()
    assert len(body) > parsing.MAX_PASTE_BYTES

    error = _error(body, pasted=True)
    assert (error.message_key, error.status_code) == ("impex.errors.file_too_large", 413)
    # The same bytes as an upload are fine — only the paste path is capped this tightly.
    assert len(parse_source(body).rows) == MAX_IMPORT_ROWS


def test_an_empty_or_header_only_source_is_an_empty_file() -> None:
    assert _error(b"").message_key == "impex.errors.empty_file"
    assert _error(b"   \n").message_key == "impex.errors.empty_file"
    assert _error(b"name,city\n").message_key == "impex.errors.empty_file"
    assert _error(_xlsx([["name", "city"]])).message_key == "impex.errors.empty_file"
